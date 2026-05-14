import streamlit as st
import pandas as pd
import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import io
import base64
 
st.set_page_config(page_title="CTS Daily Payment Sheet", page_icon="💳", layout="centered")
 
# Embedded logo
LOGO_B64 = "/9j/4AAQSkZJRgABAQAAAQABAAD/4gHYSUNDX1BST0ZJTEUAAQEAAAHIAAAAAAQwAABtbnRyUkdCIFhZWiAH4AABAAEAAAAAAABhY3NwAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAQAA9tYAAQAAAADTLQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAlkZXNjAAAA8AAAACRyWFlaAAABFAAAABRnWFlaAAABKAAAABRiWFlaAAABPAAAABR3dHB0AAABUAAAABRyVFJDAAABZAAAAChnVFJDAAABZAAAAChiVFJDAAABZAAAAChjcHJ0AAABjAAAADxtbHVjAAAAAAAAAAEAAAAMZW5VUwAAAAgAAAAcAHMAUgBHAEJYWVogAAAAAAAAb6IAADj1AAADkFhZWiAAAAAAAABimQAAt4UAABjaWFlaIAAAAAAAACSgAAAPhAAAts9YWVogAAAAAAAA9tYAAQAAAADTLXBhcmEAAAAAAAQAAAACZmYAAPKnAAANWQAAE9AAAApbAAAAAAAAAABtbHVjAAAAAAAAAAEAAAAMZW5VUwAAACAAAAAcAEcAbwBvAGcAbABlACAASQBuAGMALgAgADIAMAAxADb/2wBDAAUDBAQEAwUEBAQFBQUGBwwIBwcHBw8LCwkMEQ8SEhEPERETFhwXExQaFRERGCEYGh0dHx8fExciJCIeJBweHx7/2wBDAQUFBQcGBw4ICA4eFBEUHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh4eHh7/wAARCAA9AKcDASIAAhEBAxEB/8QAHAAAAwADAQEBAAAAAAAAAAAAAAUGAQMEBwII/8QANRAAAQMDAgQFAwIEBwAAAAAAAQIDBAAFERIhBhMxQRQiUWFxFTKBQpEHcqHRIzRSYpKx8P/EABgBAAMBAQAAAAAAAAAAAAAAAAACAwEE/8QALREAAQMCBAUDBAMBAAAAAAAAAQACEQMhBBIxQVFhkaHwBRRxExXB0YGx8eH/2gAMAwEAAhEDEQA/AP2XRRSG48Tw2phgQG13GcNiyz0T/Mo7CkfUawS4qlKi+qYYJ834J9RUTK4mvceUGkR7dOeUf8lFUtTjY/3LGUjHuBVLw7c03a1tzOSplZKkONKOShSSQQf2qdPEMqOyjVWrYOrRYHu0+fOuiY0UUo4ivKbUmO23HXKlyl6GGEHBWe+/YD1qr3tY3M5Qp03VXBjBJKb0Umt9yuZecRdbWiE0lvmc9MgLb+DsCDW63X+z3CR4aHObcdxkJwRqHqnI8w+M0grMMXieNv7VHYeo2bSBuLjqLJnRSYcVcP8ANQ19UZClqKRkKABBIwTjA6d63QL/AGeep5MSc26WU63AARhPruNx7igVqZMBw6odhqzRJYehTOik8biiwSX2mWbk0pbv2eVQCj6ZIxn261vF7tRuBt4mt+KCigtYOQQNR7enetFamdHDqsOHrCxYeOh0TGilUXiGzSpnhI89tx4khIwQFEdQlRGFfg0rtnGNvU5KaukliK41McYbSAo5SlWAVdcZ98ClOIpCJcLp24Ou4GGG3K9+SqaKX3K92u3KQmXLShTg1JSElRI9cJBOPeviVf7NGZjvP3BlDUgFTK8khYHXBFMarBMuFlNtCq6CGm+ltUzopZH4gsz9vdntT2jGZOHFqynQfQgjIriufEEZ2yyJdsnMNqZKQpcllYSnJHUYB3rDWpgTITNwtYuy5TrGh1VBRXFGucJ3xCBJQVxNpAwRo2zvntiuZ/iOysssvOTkhL6NbeEKJKf9WAMge5pjVYBJIStoVXGA09P5TaitcZ9mSwh+O6h1pYylaDkEexopwZUyCDBWw1Dmx2yXfb5bZMJkvOAS2HtOFjWCDuN8BQz+auKm+IYjkjiS3CHMdhySy6FuthJPLGMAggj7sf1qFej9QC0xsunC4n6BMuyzab2MiNOn8pFaJLDUC2tvMSo0Pw6kuJhIWC5JSrSoKKN8nGd9jmnfCUhlN7u0FrypUpEpKNQJQVjzpOD1Chn81OwoN05lwjTJT8iEZbgUIjyI/NcONWcnOPYHGc9adcFvrgylWV23mONBdYcCUgrQD+rSSCffvUsPhKwhzhEfE6Rttv8AhVxfqWFcXMYSS75jWdwL7flV1T/FVrmyZlvultDa5cBaiGnDhLqFABSc9jtsaoBS69XP6YgOKhSpCNKlLUyjIQB67/8AsV1PpCqMq46eIOHd9QeTYpRcGb5fbXNgSbc3bW3WdLalSA4orzn9Ixp/rXPHtl3m3Gzrl25mA1bMlS0uhZdOnGEgdB813q4nbFrcuJt0xDCUpUlS0gBwE423p2xIjvKKG3m1rT9yUqBI+ak7BXzOJ7XgzwVqfqYgtpgDhraRBiTuOMqNVw5NVwM/bTEaMxyWXdORuOaDnP8AL/amF3tr4vzlxbYSmKm1OsqUCAdWcgY+KokyI6nuSHmi4OqNY1ftQp6MoloutEnKSnUM9Nxil9m0AAbR2/1U+5VCSTFye8T/AFZeeWWDdbtwfa7c1bmWo/MS74zmjYBZOyeurtVHb7JIFx4hdfbS348pSy6CCdOgjPtuafRHYenkxFsYRtobUPL+B0rIlxCtKBIZKldBzBk/FLSwQYBJJI/Ufkpq/qjqhdlAAJPcg8eQ5clHWew3BAt8G4wnlNQnErS+iYOXlO6VJQBq+Qf3rYjh+aOG+IIaorZkzZLzjIyPMCcp37VYhxpQVhaToOFYPQ+9ZQpDiApCgpJ3BByDQMEwCPNIWu9Tqudmgag76gzx46qNk2a7RrqxcY7DkgKhIjuNtPhtaFJx0J2Ir6a4debl2BTMPQzFcedkIW6F8tSwDsTjO/oKssUYFb7Rkz5qD+Fn3GrAEDSN+BHHYGyj5VikLmX9bluD7ExxlbKEPBsqKUjJB7EEd+tck2z8RTbBdIbqXVJe5XhWpDyFuDCgVZUNsfJNXeBRgVhwbDNzv3n9rWepVWxYWjjtHPkJUhxPY7jIuinbYlKGbgylicdWClIUPMPU6cprXPsdwh35ybbo7j8Z2OhrlMvpaU3oGAPNsU1Z4FGBWnCMJJvOvnUpWeoVWtDYBAEfOmvxAj4S3huCLdaGYoZLOkZLZc16STkjVgZ3PpRTKiuhrQ1oaNlxveXuLnalBqdvRkWu8KvhaVIjeH5LiUjKmQCTqA7gk7+lURqIuVreufE9yCpRJjttuNMvAraUCDkFOR3T/WsfVNKCBMrlxE5RGuyIDtvlcOQ4ip8JTz0hL0gKeTqGV61DBOc9qdWNDL9+uEtkI5LKERmdI2AA1Kx+SP2qejtfVIMUPupbRIbK/CwozaNCAcHK1Hpn0xT3hCK1bpFztjBVy2HkqRqOThSEn/sGmGLFSwGv+rmotOZs6f8ACqGuS9oW5ZprbaSpa460pSBkklJwK664bvdItsS34hSit1WlttCCpaz7AbmguDbld74ymVP3SHKc/h0zDRGdVIDLQLQQdWQRkYra/azF4khLtsXw7ZjOIccbbwkHHl1Y759ad2u4s3BtxTbT7RbVpUl5pTagfgiuzUnHUVRteRbn3XMMOx0GeHZQtptr2IcSTGmMSmHgtTyIoxkEkku53Bplarctt++TlW8OSi+vw/NR9w09s9juNutVGpPqK57jJcjRFvMRlynE40tIIBVv70z8QSCSsbhGMEzp+lJWiPKVfLRKEB2OEpWmRph8lKco+33Ge5pbbo7cnhl6MzZ3nprz6uVIS1kA6uuv9OPT+9ejJWCgFXlJHQ9RWmFFjQWOTGbS03qJ05PUnJ60wxG8KfstptfvH6UjfGJ8O5CJEOr6w0ll1Q/StOy1f8c1ZRWW48ZthpIShtASkDsAMVzNW23sTVzkMJTIXnKyonr1xk7fiu0EHoc0j6mYALoo0cjiTus0UUVNdCKKKKEIooooQiiiihCKQXRqbC4h+qxoS5bDkYMvIbUAtJCiQQD164xT+ile3Mlc3MF59EQ/ItAgIs0xxbbroTrjJThClZHnX9pwewNUvDsK4tXGXOmtMx0vtNtpZQsrKdGcEnGO9O8AGs1GnQDSDOikyhlgk6Iqe4wagq8I7KXMjuNKUpqTHQVco7Zzjsdv2qhrFVqNzthUe3MIUG89fJ1jnNJXIlxm32+W9yC2481+sBIxn8dd6wIyFW66vW5ag0YRQphqIttKl9iMn7huNvWr2sYFc/tpMkqHtpMkqReswa4RjyYTaxOYS3KBJJUpYAyD8jIx8Vxz25M/h69XlbLyHZWhuO0UnUltKh29zk1eY2rGBTHDNOmkeFMcODbl4VFXNqOOIZS77FkPxVMIEPS2paUnHmA09FZrTdGcXNhcth0WzwiUxxIjreCFdwoA5Cvc1dgCjA9KU4YGb+c1hw4O6hJUPk222GVMW8G0uFCJMJxTagSMBQzkEds1T8KLdXY2C7B8Erf/AAgCABk7gHcZ600wM9KyOlUp0Mjs0pqdEMdIRRRRV1dFFFFCEUUUUIRRRRQhf//Z"
 
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap');
    html, body, [class*="css"] { font-family: 'Inter', sans-serif; }
 
    .header-block {
        background: #ffffff;
        border-radius: 16px;
        padding: 24px 36px;
        margin-bottom: 28px;
        border-top: 5px solid #e9a892;
        box-shadow: 0 2px 12px rgba(0,0,0,0.08);
        display: flex;
        align-items: center;
        gap: 20px;
    }
    .header-text h1 {
        font-size: 1.4rem; font-weight: 700;
        margin: 0 0 4px 0; color: #d97a5e;
    }
    .header-text p {
        font-size: 0.88rem; color: #888; margin: 0;
    }
 
    .instructions {
        background: #ffffff; border-radius: 12px;
        padding: 20px 24px; border-left: 4px solid #e9a892;
        margin-bottom: 24px; box-shadow: 0 2px 8px rgba(0,0,0,0.05);
    }
    .instructions h4 {
        margin: 0 0 10px 0; color: #d97a5e; font-size: 0.9rem;
        font-weight: 700; text-transform: uppercase; letter-spacing: 0.06em;
    }
    .instructions ol {
        margin: 0; padding-left: 18px; color: #555;
        font-size: 0.92rem; line-height: 1.9;
    }
 
    .result-box {
        background: #f0faf5; border: 1px solid #cee2bf;
        border-radius: 12px; padding: 20px 24px;
        margin-top: 20px; text-align: center;
    }
    .result-box h3 { color: #4a8a5e; margin: 0 0 6px 0; font-size: 1.05rem; }
    .result-box p  { color: #666; margin: 0; font-size: 0.88rem; }
 
    .stat-row { display: flex; gap: 12px; margin: 16px 0; justify-content: center; flex-wrap: wrap; }
    .stat-card {
        background: white; border-radius: 10px; padding: 14px 20px;
        text-align: center; border: 1px solid #e8e8e8;
        min-width: 110px; box-shadow: 0 1px 4px rgba(0,0,0,0.06);
    }
    .stat-card .number { font-size: 1.5rem; font-weight: 700; color: #d97a5e; }
    .stat-card .label  {
        font-size: 0.75rem; color: #999;
        text-transform: uppercase; letter-spacing: 0.05em; margin-top: 3px;
    }
 
    div[data-testid="stFileUploader"] {
        border: 2px dashed #e9a892;
        border-radius: 12px; padding: 12px; background: #fffaf9;
    }
 
    .stButton > button {
        background: linear-gradient(135deg, #e9a892, #d97a5e) !important;
        color: white !important; border: none !important;
        border-radius: 10px !important; padding: 12px 32px !important;
        font-size: 1rem !important; font-weight: 600 !important;
        width: 100% !important;
    }
    .stButton > button:hover {
        background: linear-gradient(135deg, #d97a5e, #c4654a) !important;
    }
    .stDownloadButton > button {
        background: linear-gradient(135deg, #5bb8b4, #3a9e9a) !important;
        color: white !important; border: none !important;
        border-radius: 10px !important; padding: 12px 32px !important;
        font-size: 1rem !important; font-weight: 600 !important;
        width: 100% !important; margin-top: 10px !important;
    }
    footer { visibility: hidden; }
    .block-container { padding-top: 2rem; }
</style>
""", unsafe_allow_html=True)
 
# Header with logo
st.markdown(f"""
<div class="header-block">
    <img src="data:image/png;base64,{LOGO_B64}" style="height:60px;">
    <div class="header-text">
        <h1>Daily Payment Sheet</h1>
        <p>Front Desk Payment Processing Tool</p>
    </div>
</div>
""", unsafe_allow_html=True)
 
st.markdown("""
<div class="instructions">
    <h4>How to use</h4>
    <ol>
        <li>Run your daily payment report in Practice Pro and save the <strong>.xls</strong> file</li>
        <li>Upload the file below</li>
        <li>Click <strong>Generate Payment Sheet</strong></li>
        <li>Download the finished Excel file — ready to print or use at the front desk</li>
    </ol>
</div>
""", unsafe_allow_html=True)
 
 
def process_payment_sheet(uploaded_file):
    df = pd.read_excel(uploaded_file, engine="xlrd", header=None)
 
    header_row = None
    for i, row in df.iterrows():
        if "Appt. Time" in str(list(row.values)):
            header_row = i
            break
    if header_row is None:
        raise ValueError("Could not find Appt. Time header. Please check you uploaded the correct Practice Pro report.")
 
    cols = df.iloc[header_row].values
    data = df.iloc[header_row + 1:].copy()
    data.columns = range(len(data.columns))
    col_map = {str(val).strip(): idx for idx, val in enumerate(cols)}
 
    appt_col = col_map.get("Appt. Time", 1)
    name_col = col_map.get("Patient Name", 2)
    fee_col  = col_map.get("Visit Fee", 6)
 
    data = data[data[appt_col].apply(lambda x: isinstance(x, datetime.datetime))].copy()
    data["visit_fee"]    = pd.to_numeric(data[fee_col], errors="coerce").fillna(0)
    data["patient_name"] = data[name_col].astype(str).str.strip()
    data["appt_time"]    = data[appt_col]
    data = data[data["visit_fee"] > 0]
 
    grouped = (
        data.groupby("patient_name", sort=False)
        .agg(visit_fee=("visit_fee", "sum"), first_appt=("appt_time", "min"))
        .reset_index()
        .sort_values("first_appt")
        .reset_index(drop=True)
    )
 
    report_date  = data["appt_time"].min().date()
    total_amount = grouped["visit_fee"].sum()
 
    wb = Workbook()
    ws = wb.active
    ws.title = "Daily Payment Sheet"
 
    ws.merge_cells("A1:G1")
    ws["A1"].value     = "Daily Payment Sheet Report"
    ws["A1"].font      = Font(name="Arial", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
 
    ws.merge_cells("A2:G2")
    ws["A2"].value     = report_date.strftime("%A, %B %d, %Y")
    ws["A2"].font      = Font(name="Arial", size=11)
    ws["A2"].alignment = Alignment(horizontal="center")
 
    ws.append([])
 
    thin_black     = Side(style="thin",   color="000000")
    med_black      = Side(style="medium", color="000000")
    full_border    = Border(top=thin_black, bottom=thin_black, left=thin_black, right=thin_black)
    header_border  = Border(top=thin_black, bottom=med_black,  left=thin_black, right=thin_black)
    summary_border = Border(top=med_black,  bottom=med_black,  left=thin_black, right=thin_black)
 
    headers     = ["Patient Name", "Visit Fee", "Cash", "Credit", "Check", "Notes", "Posted in PPro"]
    header_fill = PatternFill("solid", start_color="D97A5E", end_color="D97A5E")
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
 
    ws.append(headers)
    for col_idx in range(1, 8):
        cell           = ws.cell(row=4, column=col_idx)
        cell.fill      = header_fill
        cell.font      = header_font
        cell.border    = header_border
        cell.alignment = Alignment(horizontal="left" if col_idx == 1 else "center")
 
    for i, row in grouped.iterrows():
        row_num = 5 + i
 
        name_cell      = ws.cell(row=row_num, column=1, value=row["patient_name"])
        name_cell.font = Font(name="Arial", size=10)
 
        fee_cell               = ws.cell(row=row_num, column=2, value=row["visit_fee"])
        fee_cell.number_format = "$#,##0.00"
        fee_cell.font          = Font(name="Arial", size=10)
        fee_cell.alignment     = Alignment(horizontal="right")
 
        for col_idx in range(3, 7):
            ws.cell(row=row_num, column=col_idx, value="")
 
        posted_cell           = ws.cell(row=row_num, column=7, value="")
        posted_cell.alignment = Alignment(horizontal="center", vertical="center")
        posted_cell.fill      = PatternFill("solid", start_color="E8F5E9", end_color="E8F5E9")
 
        if i % 2 == 1:
            for col_idx in range(1, 7):
                ws.cell(row=row_num, column=col_idx).fill = PatternFill("solid", start_color="FDF5F3", end_color="FDF5F3")
            ws.cell(row=row_num, column=7).fill = PatternFill("solid", start_color="DCEDC8", end_color="DCEDC8")
 
        for col_idx in range(1, 8):
            ws.cell(row=row_num, column=col_idx).border = full_border
 
    summary_row  = 5 + len(grouped)
    summary_fill = PatternFill("solid", start_color="D97A5E", end_color="D97A5E")
    summary_font = Font(name="Arial", bold=True, color="FFFFFF", size=10)
 
    ws.cell(row=summary_row, column=1, value="{} Patients".format(len(grouped)))
    fee_sum = ws.cell(row=summary_row, column=2, value=total_amount)
    fee_sum.number_format = "$#,##0.00"
    for col_idx in range(1, 8):
        c        = ws.cell(row=summary_row, column=col_idx)
        c.fill   = summary_fill
        c.font   = summary_font
        c.border = summary_border
 
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 30
    ws.column_dimensions["G"].width = 14
 
    ws.freeze_panes           = "A5"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToPage   = True
    ws.page_setup.fitToWidth  = 1
    ws.print_title_rows       = "4:4"
 
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer, len(grouped), total_amount, report_date
 
 
uploaded_file = st.file_uploader(
    "Upload your Practice Pro daily payment report (.xls)",
    type=["xls"],
    help="Export the Daily Payment Slip report from Practice Pro and upload it here"
)
 
if uploaded_file:
    if st.button("Generate Payment Sheet"):
        with st.spinner("Processing your report..."):
            try:
                buffer, patient_count, total_amount, report_date = process_payment_sheet(uploaded_file)
                filename = "Daily_Payment_Sheet_{}.xlsx".format(report_date.strftime("%Y-%m-%d"))
 
                st.markdown("""
                <div class="result-box">
                    <h3>✅ Payment Sheet Ready!</h3>
                    <p>Your daily sheet has been generated and is ready to download.</p>
                </div>
                """, unsafe_allow_html=True)
 
                st.markdown("""
                <div class="stat-row">
                    <div class="stat-card"><div class="number">{}</div><div class="label">Patients</div></div>
                    <div class="stat-card"><div class="number">${:,.2f}</div><div class="label">Total to Collect</div></div>
                    <div class="stat-card"><div class="number">{}</div><div class="label">Report Date</div></div>
                </div>
                """.format(patient_count, total_amount, report_date.strftime("%b %d")), unsafe_allow_html=True)
 
                st.download_button(
                    label="⬇️  Download Daily Payment Sheet",
                    data=buffer,
                    file_name=filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
 
            except Exception as e:
                st.error("Error processing file: {}".format(str(e)))
                st.info("Please make sure you uploaded the correct Practice Pro Daily Payment Slip report (.xls format).")
else:
    st.markdown("""
    <div style="text-align:center; padding: 32px; color: #ccc; font-size: 0.9rem;">
        ⬆️ Upload your .xls file above to get started
    </div>
    """, unsafe_allow_html=True)
 
st.markdown("---")
st.markdown(
    "<div style=\'text-align:center; color:#ccc; font-size:0.78rem;\'>Comprehensive Therapy Services · Daily Payment Processing</div>",
    unsafe_allow_html=True
)
 
